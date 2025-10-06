import datetime
import logging
from io import BytesIO

from aiogram import Bot, F, Router, types
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, ChatPermissions
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from dbase.orm_query import (
    change_restrict_word,
    orm_add_update_meter,
    orm_add_word,
    orm_del_user,
    orm_del_word_obj,
    orm_get_all_meters_to_month,
    orm_get_unconfirmed_user_last,
    orm_get_user_apartment,
    orm_get_user_meters_last,
    orm_get_user_tele,
    orm_get_users_confirm,
    orm_get_users_to_apart,
    orm_get_word_obj,
    orm_get_words,
)
from filters.chat_types import ChatTypeFilter, IsAdmin
from filters.data_filter import validate_apart, validate_data_meter, validate_porch
from handlers.const import PORCH_APART
from handlers.states import ChangeMeter, ChangeWords, PorchMessage, SetApart
from kbds.kbds import (
    btns,
    btns_admin,
    btns_cnl,
    btns_edit_del_new,
    btns_yes_no,
    get_user_main_btns,
)

logger = logging.getLogger(__name__)

user_private_admin_router = Router()
user_private_admin_router.message.filter(ChatTypeFilter(["private"]), IsAdmin())
user_private_admin_router.callback_query.filter(IsAdmin())


@user_private_admin_router.message(Command("menu"))
async def nemu_cmd(message: types.Message, state: FSMContext):
    await start_cmd(message, state)


@user_private_admin_router.message(Command("about"))
async def about_cmd(message: types.Message, state: FSMContext):
    text_mgs = (
        f"Приветствую Вас, {message.from_user.username}!\n"
        "Это бот для жителей дома №6 мкр. Рождественский."
        "\nВы являетесь администратором."
        "\nДля начала работы отправьте команду /start"
    )
    await message.answer(text_mgs)
    await start_cmd(message, state)


@user_private_admin_router.message(CommandStart())
async def start_cmd(message: types.Message, state: FSMContext):
    await state.clear()
    text_mgs = "Добро пожаловать , Администратор"
    await message.answer(text_mgs, reply_markup=get_user_main_btns(btns_admin))


@user_private_admin_router.callback_query(F.data == "cancel")
async def cancel_cmd(
    callback: types.CallbackQuery,
    state: FSMContext,
):
    await callback.answer()
    await callback.bot.delete_message(
        chat_id=callback.message.chat.id, message_id=callback.message.message_id
    )
    await start_cmd(callback.message, state)


@user_private_admin_router.callback_query(F.data == "restrict_words")
async def restrict_words_cmd(callback: types.CallbackQuery):
    await callback.message.edit_text(
        "Ругательства\nЧто будем делать?",
        reply_markup=get_user_main_btns(btns_edit_del_new),
    )


@user_private_admin_router.callback_query(F.data == "edit_word")
async def edit_word_cmd(
    callback: types.CallbackQuery, session: AsyncSession, state: FSMContext
):
    await get_word_cmd(callback, session)
    await callback.message.edit_text(
        f"Введите изменяемое слово.\n" f"Полный список слов Вам выслан."
    )
    await state.set_state(ChangeWords.edit_word)


@user_private_admin_router.message(StateFilter(ChangeWords.edit_word))
async def state_enter_edit_word_cmd(
    message: types.Message, session: AsyncSession, state: FSMContext
):
    word = message.text.lower().strip()
    word_obj = await orm_get_word_obj(session, word)
    if not word_obj:
        btns = {"Ввести заново": "edit_word", "Отмена": "cancel"}
        await message.answer("Слово не найдено.", reply_markup=get_user_main_btns(btns))
        return
    await state.update_data(old_word=word_obj)
    await message.answer(f'Принято слово для изменения "{word}"')
    await message.answer("Введите слово, на которое нужно заменить:")
    await state.set_state(ChangeWords.input_word)


@user_private_admin_router.message(StateFilter(ChangeWords.input_word))
async def input_word_cmd(
    message: types.Message,
    session: AsyncSession,
    state: FSMContext,
):
    new_word = message.text.lower().strip()
    if not new_word:
        await message.answer("Пустое слово. Введите заново.")
        return

    data = await state.get_data()
    old_word = data.get("old_word").word
    change = await change_restrict_word(session, old_word=old_word, new_word=new_word)
    if change:
        await message.answer(f'Слово "{old_word}" успешно заменено на "{new_word}".')
        await state.clear()
        await start_cmd(message, state)
    else:
        await message.answer("Ошибка замены.")
        logger.error(f"Ошибка замены {change}.")


@user_private_admin_router.callback_query(F.data == "del_word")
async def del_word_cmd(
    callback: types.CallbackQuery, session: AsyncSession, state: FSMContext
):
    await get_word_cmd(callback, session)
    await callback.message.edit_text(
        f"Введите удаляемое слово.\n" f"Полный список слов Вам выслан."
    )
    await state.set_state(ChangeWords.delete_word)


@user_private_admin_router.message(StateFilter(ChangeWords.delete_word))
async def delete_word_cmd(
    message: types.Message,
    session: AsyncSession,
    state: FSMContext,
):
    word = message.text.lower().strip()
    word_obj = await orm_get_word_obj(session, word)
    if not word_obj:
        await message.answer("Слово не найдено. Введите заново.")
        return
    try:
        await orm_del_word_obj(session, word_obj)
        await message.answer(f'Слово "{word}" успешно удалено.')
    except Exception as e:
        await message.answer(f"При удалении возникла ошибка.")
        logger.error(f"При удалении слова {word} возникла ошибка: {e}")
    await state.clear()
    logger.info(
        f"Успешное удаление слова {word} " f"админом {message.from_user.username}"
    )
    await start_cmd(message, state)


@user_private_admin_router.callback_query(F.data == "add_word")
async def add_word_cmd(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text(f"Введите запрещённое слово.")
    await state.set_state(ChangeWords.add_word)


@user_private_admin_router.message(StateFilter(ChangeWords.add_word))
async def add_word_cmd(
    message: types.Message,
    session: AsyncSession,
    state: FSMContext,
):
    word = message.text.lower().strip()
    word_obj = await orm_get_word_obj(session, word)
    if word_obj:
        btn = {"Отмена": "cancel"}
        await message.answer(
            f'Слово "{word}" уже есть в базе.\nВведите другое слово.',
            reply_markup=get_user_main_btns(btn),
        )
        return
    try:
        await orm_add_word(session, word)
        await message.answer(f'Слово "{word}" успешно добавлено.')
    except:
        await message.answer(f'Ошибка добавления слова "{word}"')
    finally:
        await state.clear()
        await start_cmd(message, state)


#################### удаление изменение запрещённых слов меню  ##################


@user_private_admin_router.callback_query(F.data.startswith("block_user_"))
async def block_user_cmd(callback: types.CallbackQuery, bot: Bot):
    await callback.answer()
    data_callback = callback.data.split("_")
    block_user_id = data_callback[-2]
    block_chat_id = data_callback[-1]
    try:
        await bot.restrict_chat_member(
            chat_id=int(block_chat_id),
            user_id=int(block_user_id),
            permissions=ChatPermissions(can_send_messages=False),
        )
        await callback.answer(
            "Пользователю заблокирована отправка сообщений в чате! Не удален!",
            show_alert=True,
        )
        await callback.message.edit_text(
            callback.message.text + "\n\n✅ Пользователь успешно ограничен."
        )

    except Exception as e:
        await callback.message.edit_text(
            f"Ошибка: блокировки\n Сообщите разработчику\nчат:{block_chat_id}, user:{block_user_id}",
            show_alert=True,
        )
        logger.error(
            f"Ошибка блокировки: {e} чат:{block_chat_id}, user:{block_user_id}"
        )


async def generate_excel_in_memory_words(
    session: AsyncSession,
):
    """Создаёт Excel-файл в памяти"""
    workbook = Workbook()
    sheet = workbook.active
    table_name = "Запрещённые слова"
    sheet.append([table_name])
    words = await orm_get_words(session)
    max_ln = len(table_name)

    for word in words:
        sheet.append([word])
        max_ln = max(max_ln, len(word))

    column_letter = get_column_letter(1)
    sheet.column_dimensions[column_letter].width = max_ln + 2
    sheet[column_letter][0].alignment = Alignment(
        horizontal="center", vertical="center"
    )

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook


@user_private_admin_router.callback_query(F.data == "get_words")
async def get_word_cmd(callback: types.CallbackQuery, session: AsyncSession):
    await callback.answer()
    virtual_workbook = await generate_excel_in_memory_words(session)
    filename = (
        "Запрещенные слова " f"{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx"
    )
    document = BufferedInputFile(file=virtual_workbook.getvalue(), filename=filename)
    await callback.bot.send_document(
        chat_id=callback.message.chat.id,
        document=document,
        caption="Список запрещённых слов готов!",
    )


@user_private_admin_router.callback_query(F.data == "confirm_user")
async def confirm_user_cmd(callback: types.CallbackQuery, session: AsyncSession):
    user = await orm_get_unconfirmed_user_last(session)
    if not user:
        await callback.answer("Нет неподтверждённых пользователей.", show_alert=True)
        return
    user_info = (
        f"👤 <b>Имя:</b> {user.name}\n"
        f"👤 <b>Квартира:</b> {user.apartment}\n"
        f"📞 <b>Телефон:</b> {user.phone}\n"
        f"🆔 <b>Telegram ID:</b> {user.tele_id}\n"
        "📅 <b>Дата регистрации:</b> "
        f"{user.created.strftime('%d.%m.%Y %H:%M') if user.created else '—'}"
    )
    await callback.message.edit_text(
        user_info,
        reply_markup=get_user_main_btns(
            {
                "Подтвердить": f"conf_user_{user.tele_id}",
                "Удалить": f"del_user_{user.tele_id}",
            }
        ),
        parse_mode="HTML",
    )
    await callback.answer()


@user_private_admin_router.callback_query(F.data.startswith("conf_user"))
async def conf_user_cmd(
    callback: types.CallbackQuery, session: AsyncSession, bot: Bot, state: FSMContext
):
    tele_id = int(callback.data.split("_")[-1])
    user = await orm_get_user_tele(session, tele_id)
    user.confirmed = True
    await session.commit()
    await bot.send_message(
        chat_id=user.tele_id, text=f"✅ Вас подтвердили! Добро пожаловать, {user.name}."
    )
    await callback.answer(
        f"Пользователь {user.name} - кв {user.apartment} подтвержден.", show_alert=True
    )
    await start_cmd(callback.message, state)


@user_private_admin_router.callback_query(F.data.startswith("del_user"))
async def del_user_cmd(callback: types.CallbackQuery, session: AsyncSession, bot: Bot):
    tele_id = int(callback.data.split("_")[-1])
    user = await orm_get_user_tele(session, tele_id)
    await orm_del_user(session, user.tele_id)
    await callback.answer(
        f"Пользователь {user.name} - кв {user.apartment} удалён.", show_alert=True
    )
    await bot.send_message(
        chat_id=user.tele_id,
        text=(
            f"❌ Вы не прошли проверку, {user.name}. Похоже Вы не из наших."
            "\nПрощайте."
        ),
    )


@user_private_admin_router.callback_query(F.data == "edit_meter")
async def edit_meter_cmd(callback: types.CallbackQuery, bot: Bot, state: FSMContext):
    user_tele_id = callback.from_user.id
    await bot.send_message(user_tele_id, text="Введите номер квартиры")
    await state.set_state(ChangeMeter.apartment)
    await callback.answer()


@user_private_admin_router.callback_query(F.data.startswith("msg_porch"))
async def msg_porch_cmd(callback: types.CallbackQuery, bot: Bot, state: FSMContext):
    user_tele_id = callback.from_user.id
    await bot.send_message(user_tele_id, text="Введите номер подъезда")
    await state.set_state(PorchMessage.porch)
    await callback.answer()


@user_private_admin_router.message(ChangeMeter.apartment, F.text)
async def input_apart(message: types.Message, session: AsyncSession, state: FSMContext):
    if await validate_apart(message):
        user = await orm_get_user_apartment(session, message.text)
        if user is None:
            await message.answer("Квартира не найдена")
            await start_cmd(message, state)
            return

        await state.update_data(apartment=message.text)
        meter = await orm_get_user_meters_last(session, user.tele_id)
        user_info = f"Текущие показания кв {user.apartment}:\n"
        last_meter_data = await orm_get_user_meters_last(session, user_id=user.tele_id)
        if meter:
            user_info += (
                f"🚰 <b>Счётчики горячей воды:</b> кухня - "
                f"{meter.water_hot_kitchen if meter.water_hot_kitchen else 'Не найдено'}    СУ - {meter.water_hot_bath if meter.water_hot_bath else 'Не найдено'}\n"
                f"🚰 <b>Счётчики холодной воды:</b> кухня - "
                f"{meter.water_cold_kitchen if meter.water_cold_kitchen else 'Не найдено'}    СУ - {meter.water_cold_bath if meter.water_cold_bath else 'Не найдено'}\n"
                f"\nПоследние показания на  {last_meter_data.updated}"
            )
        else:
            user_info = "Показания не обнаружены"
        await message.answer(
            user_info, parse_mode="HTML", reply_markup=get_user_main_btns(btns)
        )


@user_private_admin_router.message(PorchMessage.porch, F.text)
async def input_porch(message: types.Message, state: FSMContext):
    if await validate_porch(message):
        await state.update_data(porch=message.text)
        await message.answer("Введите сообщение")
        await state.set_state(PorchMessage.text)


@user_private_admin_router.message(PorchMessage.text, F.text)
async def input_text(message: types.Message, state: FSMContext):
    await state.update_data(text=message.text)
    data = await state.get_data()
    await message.answer(
        f'В {data.get("porch")} подъезд будет отправлено'
        f' сообщение:\n{data.get("text")}',
        reply_markup=get_user_main_btns(btns_yes_no),
    )
    await state.set_state(PorchMessage.confirm)


@user_private_admin_router.callback_query(PorchMessage.confirm)
async def send_msg_porch(
    callback: types.CallbackQuery, session: AsyncSession, bot: Bot, state: FSMContext
):
    if callback.data == "yes":
        data = await state.get_data()
        porch = int(data.get("porch"))
        aparts = PORCH_APART[porch]
        users = await orm_get_users_to_apart(session, aparts[0], aparts[1])
        for user in users:
            try:
                await bot.send_message(user.tele_id, text=data.get("text"))
            except TelegramForbiddenError:
                await orm_del_user(session, user.tele_id)
                txt = (
                    f"Пользователь {user.tele_id} - {user.apartment} "
                    "заблокировал бота и был удалён."
                )
                await bot.send_message(callback.from_user.id, text=txt)
            except TelegramBadRequest as e:
                print(f"Неверный запрос при отправке {user.tele_id}: {e}")
            except Exception as e:
                print(f"Неизвестная ошибка при отправке {user.tele_id}: {e}")
        await callback.answer("Сообщения разосланы", show_alert=True)
    elif callback.data == "cancel":
        await state.clear()
        await callback.answer("Отправка отменена", show_alert=True)
    await start_cmd(callback.message, state)


@user_private_admin_router.callback_query(F.data.startswith("get_meter_all"))
async def get_meter_all_cmd(
    callback: types.CallbackQuery, session: AsyncSession, bot: Bot
):
    users = await orm_get_users_confirm(session)
    for user in users:
        try:
            await bot.send_message(
                user.tele_id,
                "Здравствуйте.\nПрошу Вас передать " "показания приборов учёта.",
                reply_markup=get_user_main_btns(btns),
            )
        except TelegramForbiddenError:
            print(
                f"Пользователь {user.tele_id} - " f"{user.apartment} заблокировал бота"
            )
            await orm_del_user(session, user.tele_id)
            txt = (
                f"Пользователь {user.tele_id} - {user.apartment} "
                "заблокировал бота и был удалён."
            )
            await bot.send_message(callback.from_user.id, text=txt)
        except TelegramBadRequest as e:
            print(f"Неверный запрос при отправке {user.tele_id}: {e}")
        except Exception as e:
            print(f"Неизвестная ошибка при отправке {user.tele_id}: {e}")
    await callback.answer()


@user_private_admin_router.callback_query(F.data == "get_data_apart")
async def get_data_apart_cmd(
    callback: types.CallbackQuery, bot: Bot, state: FSMContext
):
    await bot.send_message(callback.from_user.id, "Введите номер квартиры")
    await callback.answer()
    await state.set_state(SetApart.apartment)


async def show_meter_info(
    message: types.Message, session: AsyncSession, state: FSMContext, apartment: str
):
    user = await orm_get_user_apartment(session, apartment)
    if user is None:
        await message.answer("Квартира не найдена")
        await start_cmd(message, state)
        return

    await state.update_data(apartment=apartment)  # Обновляем состояние

    meter = await orm_get_user_meters_last(session, user.tele_id)

    user_info = f"Текущие показания кв {user.apartment}:\n"
    user_info += (
        f"Имя: {user.name}\nТелефон: {user.phone}\nПодтверждён: "
        f'{"Да" if user.confirmed else "Нет"}\n'
    )
    if meter:
        user_info += (
            f"🚰 <b>Счётчики горячей воды:</b> кухня - {meter.water_hot_kitchen if meter.water_hot_kitchen else 'Не найдено'}    СУ - {meter.water_hot_bath if meter.water_hot_bath else 'Не найдено'}\n"
            f"🚰 <b>Счётчики холодной воды:</b> кухня - {meter.water_cold_kitchen if meter.water_cold_kitchen else 'Не найдено'}    СУ - {meter.water_cold_bath if meter.water_cold_bath else 'Не найдено'}\n"
            f"\nПоследние показания на  {meter.updated}"
        )
    else:
        user_info = "Показания не обнаружены"

    await message.answer(
        user_info, parse_mode="HTML", reply_markup=get_user_main_btns(btns_cnl)
    )


@user_private_admin_router.message(SetApart.apartment, F.text)
async def send_info_apart(
    message: types.Message, session: AsyncSession, state: FSMContext
):
    if not await validate_apart(message):
        return

    await show_meter_info(message, session, state, message.text)


async def generate_excel_in_memory(
    session: AsyncSession,
):
    """Создаёт Excel-файл в памяти"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Квартира",
            "Горячая вода (ванна)",
            "Холодная вода (ванна)",
            "Горячая вода (кухня)",
            "Холодная вода (кухня)",
            "Дата списания",
        ]
    )
    meters = await orm_get_all_meters_to_month(session)
    print(f"Найдено записей: {len(meters)}")

    for i, meter in enumerate(meters):
        sheet.append(
            [
                meter.user.apartment,
                meter.water_hot_bath or 0,
                meter.water_cold_bath or 0,
                meter.water_hot_kitchen or 0,
                meter.water_cold_kitchen or 0,
                meter.created.strftime("%Y-%m-%d %H:%M") if meter.created else "",
            ]
        )

    for i in range(1, 7):
        column_letter = get_column_letter(i)
        length = len(str(sheet[column_letter][0].value))
        sheet.column_dimensions[column_letter].width = min(length + 3, 50)
        sheet[column_letter][0].alignment = Alignment(
            horizontal="center", vertical="center"
        )

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook


@user_private_admin_router.callback_query(F.data == "get_meter_month")
async def get_meter_month(
    callback: types.CallbackQuery,
    bot: Bot,
    session: AsyncSession,
):
    await callback.answer()
    virtual_workbook = await generate_excel_in_memory(session)
    filename = (
        f"Счетчики воды на " f"{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx"
    )
    document = BufferedInputFile(file=virtual_workbook.getvalue(), filename=filename)
    await bot.send_document(
        chat_id=callback.message.chat.id, document=document, caption="Ваш отчёт готов!"
    )


@user_private_admin_router.callback_query(F.data)
async def set_meter_cmd(
    callback_query: types.CallbackQuery, session: AsyncSession, state: FSMContext
):

    action = callback_query.data

    # Словарь: сопоставление callback_data и состояний
    state_mapping = {
        "water_hot_kitchen": ChangeMeter.water_hot_kitchen,
        "water_cold_kitchen": ChangeMeter.water_cold_kitchen,
        "water_hot_bath": ChangeMeter.water_hot_bath,
        "water_cold_bath": ChangeMeter.water_cold_bath,
    }
    data = await state.get_data()
    user = await orm_get_user_apartment(session, data["apartment"])
    meter = await orm_get_user_meters_last(session, user.tele_id)
    current_value = None
    name_meter = ""
    if action == "water_hot_kitchen":
        current_value = meter.water_hot_kitchen if meter else None
        name_meter = "Горячая вода кухня"
    elif action == "water_cold_kitchen":
        current_value = meter.water_cold_kitchen if meter else None
        name_meter = "Холодная вода кухня"
    elif action == "water_hot_bath":
        current_value = meter.water_hot_bath if meter else None
        name_meter = "Горячая вода СУ"
    elif action == "water_cold_bath":
        current_value = meter.water_cold_bath if meter else None
        name_meter = "Холодная вода СУ"

    msg = (
        f"{name_meter}\nТекущие показания - "
        f'{current_value if current_value else " не найдены."}'
        "\nВведите показания счётчика."
    )
    await callback_query.message.answer(msg)
    try:
        await state.set_state(state_mapping[action])
    except Exception as e:
        logger.error(f"Ошибка {e}")
        await start_cmd(callback_query.message, state)
    await callback_query.answer()


@user_private_admin_router.message(F.text, StateFilter("*"))
async def save_meter_cmd(
    message: types.Message, session: AsyncSession, state: FSMContext
):

    current_state = await state.get_state()
    if not current_state or not current_state.startswith("ChangeMeter"):
        return  # Игнорируем, если не в нужном состоянии

    data = await state.get_data()
    apartment = data.get("apartment")
    if not apartment:
        await message.answer("Ошибка: квартира не выбрана.")
        await state.clear()
        return

    user = await orm_get_user_apartment(session, apartment)
    if not user:
        await message.answer("Пользователь не найден.")
        await state.clear()
        return

    meter = await orm_get_user_meters_last(session, user.tele_id)

    water_hot_kitchen_data = None
    water_cold_kitchen_data = None
    water_hot_bath_data = None
    water_cold_bath_data = None

    meter_value = None

    if current_state == ChangeMeter.water_hot_kitchen.state:
        meter_value = meter.water_hot_kitchen if meter else None
        water_hot_kitchen_data = message.text
    elif current_state == ChangeMeter.water_cold_kitchen.state:
        meter_value = meter.water_cold_kitchen if meter else None
        water_cold_kitchen_data = message.text
    elif current_state == ChangeMeter.water_hot_bath.state:
        meter_value = meter.water_hot_bath if meter else None
        water_hot_bath_data = message.text
    elif current_state == ChangeMeter.water_cold_bath.state:
        meter_value = meter.water_cold_bath if meter else None
        water_cold_bath_data = message.text
    else:
        await message.answer("Неизвестное состояние.")
        return

    # Валидация
    validate = await validate_data_meter(message, state, message.text, meter_value)
    if not validate:
        return

    await orm_add_update_meter(
        session,
        user.tele_id,
        water_hot_kitchen=water_hot_kitchen_data,
        water_cold_kitchen=water_cold_kitchen_data,
        water_hot_bath=water_hot_bath_data,
        water_cold_bath=water_cold_bath_data,
    )
    await message.answer("✅ Показания сохранены.")
    await show_meter_info(message, session, state, apartment)
